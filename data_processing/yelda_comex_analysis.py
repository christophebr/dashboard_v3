#!/usr/bin/env python3
"""
Analyse de l'utilisation du chatbot Yelda (Stellair) — classeur Excel COMEX.

Périmètre :
  - URL d'origine = https://fse.stellair.fr (chatbot FSE Stellair)
  - Conversations **initiées** uniquement : nombre de messages >= 2
    (exclut les simples ouvertures du widget : parcours "welcome" seul,
    nb_messages = 1, durée quasi nulle).

Contenu du classeur :
  - Synthese                 : indicateurs clés + messages COMEX
  - Volumetrie_mensuelle     : conversations initiées par mois
  - Utilisateurs             : utilisateurs HubSpot uniques et récurrents
  - Satisfaction_LLM_mensuel : taux de satisfaction LLM par mois
  - Deflexion_ticket         : part des conversations débouchant sur un ticket
  - Top_intentions           : top intentions sur les conversations initiées
  - Methodologie             : définition du périmètre et des indicateurs

Usage :
  python3 data_processing/yelda_comex_analysis.py
  python3 data_processing/yelda_comex_analysis.py -o rapport.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict

import pandas as pd

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from data_processing.yelda_processing import (  # noqa: E402
    COL_DATE,
    COL_EVALUATION_LLM,
    COL_HUBSPOT_CONTACT_ID,
    COL_INTENTIONS,
    COL_PARCOURS,
    filter_yelda_evaluated,
    filter_yelda_stellair,
    has_ticket_created,
    load_yelda_data,
    normalize_yelda_hubspot_contact_id,
)

COL_NB_MSG = "Nombre de messages"
COL_DUREE = "Durée en seconde"

SEUIL_MESSAGES_INITIEE = 2  # >= 2 => l'utilisateur a envoyé au moins une question


def filter_conversations_initiees(df_fse: pd.DataFrame) -> pd.DataFrame:
    """Ne garde que les conversations avec au moins un message utilisateur
    (nb_messages >= 2, parcours != 'welcome' strict)."""
    if df_fse is None or df_fse.empty:
        return pd.DataFrame()
    nb = pd.to_numeric(df_fse[COL_NB_MSG], errors="coerce").fillna(0)
    return df_fse.loc[nb >= SEUIL_MESSAGES_INITIEE].copy()


def _categorise_llm(val) -> str:
    s = str(val).strip().lower()
    if "insatisfait" in s:
        return "Insatisfait"
    if "satisfait" in s:
        return "Satisfait"
    if "revoir" in s:
        return "À revoir"
    return "Non évaluée"


def build_volumetrie_mensuelle(df_init: pd.DataFrame) -> pd.DataFrame:
    d = df_init.copy()
    d["Mois"] = pd.to_datetime(d[COL_DATE]).dt.to_period("M").astype(str)
    g = d.groupby("Mois").size().rename("Conversations initiées").to_frame()
    g["Utilisateurs uniques (HubSpot ID)"] = d.groupby("Mois")[COL_HUBSPOT_CONTACT_ID].apply(
        lambda s: s.map(normalize_yelda_hubspot_contact_id).dropna().nunique()
    )
    dur = pd.to_numeric(d[COL_DUREE], errors="coerce")
    d = d.assign(_dur=dur)
    g["Durée médiane (s)"] = d.groupby("Mois")["_dur"].median().round(0).astype("Int64")
    g["Nb messages médian"] = (
        d.groupby("Mois")[COL_NB_MSG].median().round(0).astype("Int64")
    )
    g = g.reset_index()
    return g


def build_volumetrie_hebdo(df_init: pd.DataFrame) -> pd.DataFrame:
    """Volumétrie par semaine ISO : libellé 'AAAA-Sxx', conversations initiées,
    utilisateurs uniques, durée/nb messages médians."""
    if df_init is None or df_init.empty:
        return pd.DataFrame(
            columns=[
                "Semaine",
                "Date lundi",
                "Conversations initiées",
                "Utilisateurs uniques (HubSpot ID)",
                "Durée médiane (s)",
                "Nb messages médian",
            ]
        )
    d = df_init.copy()
    dt = pd.to_datetime(d[COL_DATE])
    iso = dt.dt.isocalendar()
    d["_iso_year"] = iso.year.astype(int)
    d["_iso_week"] = iso.week.astype(int)
    d["Semaine"] = (
        d["_iso_year"].astype(str) + "-S" + d["_iso_week"].astype(str).str.zfill(2)
    )
    # lundi de la semaine ISO pour tri et affichage
    d["Date lundi"] = (
        dt - pd.to_timedelta(dt.dt.weekday, unit="D")
    ).dt.normalize()
    d["_dur"] = pd.to_numeric(d[COL_DUREE], errors="coerce")
    d["_msg"] = pd.to_numeric(d[COL_NB_MSG], errors="coerce")
    g = (
        d.groupby(["Semaine", "Date lundi"])
        .agg(
            **{
                "Conversations initiées": (COL_DATE, "size"),
                "Durée médiane (s)": ("_dur", "median"),
                "Nb messages médian": ("_msg", "median"),
            }
        )
        .reset_index()
    )
    g["Utilisateurs uniques (HubSpot ID)"] = (
        d.groupby("Semaine")[COL_HUBSPOT_CONTACT_ID]
        .apply(lambda s: s.map(normalize_yelda_hubspot_contact_id).dropna().nunique())
        .reindex(g["Semaine"])
        .values
    )
    g["Durée médiane (s)"] = g["Durée médiane (s)"].round(0).astype("Int64")
    g["Nb messages médian"] = g["Nb messages médian"].round(0).astype("Int64")
    g = g.sort_values("Date lundi").reset_index(drop=True)
    g["Date lundi"] = g["Date lundi"].dt.strftime("%Y-%m-%d")
    # Ordre final de colonnes
    return g[
        [
            "Semaine",
            "Date lundi",
            "Conversations initiées",
            "Utilisateurs uniques (HubSpot ID)",
            "Durée médiane (s)",
            "Nb messages médian",
        ]
    ]


def build_utilisateurs(df_init: pd.DataFrame) -> pd.DataFrame:
    ids = df_init[COL_HUBSPOT_CONTACT_ID].map(normalize_yelda_hubspot_contact_id).dropna()
    vc = ids.value_counts()
    total_conv = len(df_init)
    sans_id = total_conv - len(ids)
    lignes = [
        ("Conversations initiées (total)", total_conv),
        ("Conversations sans ID HubSpot identifié", sans_id),
        ("Utilisateurs uniques (ID HubSpot)", int(vc.shape[0])),
        ("Utilisateurs avec 1 seule conversation", int((vc == 1).sum())),
        ("Utilisateurs avec >= 2 conversations", int((vc >= 2).sum())),
        ("Utilisateurs avec >= 3 conversations", int((vc >= 3).sum())),
        ("Utilisateurs avec >= 5 conversations", int((vc >= 5).sum())),
        ("Nb conversations par utilisateur (moyenne)", round(float(vc.mean()) if len(vc) else 0.0, 2)),
        ("Nb conversations par utilisateur (médiane)", int(vc.median()) if len(vc) else 0),
    ]
    return pd.DataFrame(lignes, columns=["Indicateur", "Valeur"])


def build_satisfaction_llm_mensuel(df_init: pd.DataFrame) -> pd.DataFrame:
    df_eval = filter_yelda_evaluated(df_init)
    if df_eval.empty:
        return pd.DataFrame(
            columns=[
                "Mois",
                "Satisfait",
                "Insatisfait",
                "À revoir",
                "Évaluées",
                "Taux satisfaction LLM (%)",
            ]
        )
    d = df_eval.copy()
    d["Mois"] = pd.to_datetime(d[COL_DATE]).dt.to_period("M").astype(str)
    d["_cat"] = d[COL_EVALUATION_LLM].map(_categorise_llm)
    pivot = (
        d.pivot_table(index="Mois", columns="_cat", values=COL_DATE, aggfunc="count", fill_value=0)
        .reset_index()
    )
    for c in ("Satisfait", "Insatisfait", "À revoir"):
        if c not in pivot.columns:
            pivot[c] = 0
    pivot["Évaluées"] = pivot["Satisfait"] + pivot["Insatisfait"] + pivot["À revoir"]
    denom = (pivot["Satisfait"] + pivot["Insatisfait"]).replace(0, pd.NA)
    pivot["Taux satisfaction LLM (%)"] = (
        100 * pivot["Satisfait"] / denom
    ).round(1)
    return pivot[
        ["Mois", "Satisfait", "Insatisfait", "À revoir", "Évaluées", "Taux satisfaction LLM (%)"]
    ]


def build_deflexion_ticket(df_init: pd.DataFrame) -> pd.DataFrame:
    d = df_init.copy()
    d["Mois"] = pd.to_datetime(d[COL_DATE]).dt.to_period("M").astype(str)
    d["_ticket"] = d[COL_PARCOURS].apply(has_ticket_created)
    g = d.groupby("Mois").agg(
        Conversations_initiees=("_ticket", "size"),
        Tickets_crees=("_ticket", "sum"),
    )
    g["Part avec ticket (%)"] = (100 * g["Tickets_crees"] / g["Conversations_initiees"]).round(1)
    g["Part sans ticket / déflexion (%)"] = (100 - g["Part avec ticket (%)"]).round(1)
    g = g.reset_index().rename(
        columns={
            "Conversations_initiees": "Conversations initiées",
            "Tickets_crees": "Tickets créés",
        }
    )
    total = pd.DataFrame(
        [
            {
                "Mois": "Total",
                "Conversations initiées": int(g["Conversations initiées"].sum()),
                "Tickets créés": int(g["Tickets créés"].sum()),
                "Part avec ticket (%)": round(
                    100 * g["Tickets créés"].sum() / g["Conversations initiées"].sum(), 1
                ),
                "Part sans ticket / déflexion (%)": round(
                    100 - (100 * g["Tickets créés"].sum() / g["Conversations initiées"].sum()), 1
                ),
            }
        ]
    )
    return pd.concat([g, total], ignore_index=True)


def build_top_intentions(df_init: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    if COL_INTENTIONS not in df_init.columns:
        return pd.DataFrame(columns=["Intention", "Occurrences", "Part des conv. initiées (%)"])
    tokens = []
    for v in df_init[COL_INTENTIONS].dropna().astype(str):
        for t in v.replace(",", ";").split(";"):
            t = t.strip()
            if t:
                tokens.append(t)
    s = pd.Series(tokens).value_counts().head(top_n)
    total = len(df_init)
    df = s.rename_axis("Intention").reset_index(name="Occurrences")
    df["Part des conv. initiées (%)"] = (100 * df["Occurrences"] / total).round(1) if total else 0
    return df


def build_synthese(df_fse: pd.DataFrame, df_init: pd.DataFrame) -> pd.DataFrame:
    dur = pd.to_numeric(df_init[COL_DUREE], errors="coerce")
    msg = pd.to_numeric(df_init[COL_NB_MSG], errors="coerce")
    ids = df_init[COL_HUBSPOT_CONTACT_ID].map(normalize_yelda_hubspot_contact_id).dropna()
    vc = ids.value_counts()
    nb_tickets = int(df_init[COL_PARCOURS].apply(has_ticket_created).sum())
    df_eval = filter_yelda_evaluated(df_init)
    # Taux sat global (hors "À revoir")
    cats = df_eval[COL_EVALUATION_LLM].map(_categorise_llm)
    nb_sat = int((cats == "Satisfait").sum())
    nb_insat = int((cats == "Insatisfait").sum())
    taux_sat = round(100 * nb_sat / (nb_sat + nb_insat), 1) if (nb_sat + nb_insat) else 0.0

    d_fse = pd.to_datetime(df_fse[COL_DATE])
    periode = f"{d_fse.min().date()} → {d_fse.max().date()}" if len(df_fse) else "—"

    lignes = [
        ("Période couverte", periode),
        ("Ouvertures du widget (FSE Stellair)", int(len(df_fse))),
        ("Conversations initiées (>= 1 question utilisateur)", int(len(df_init))),
        (
            "Taux de conversations initiées (%)",
            round(100 * len(df_init) / len(df_fse), 1) if len(df_fse) else 0.0,
        ),
        ("Utilisateurs uniques (HubSpot ID)", int(vc.shape[0])),
        ("Utilisateurs récurrents (>= 2 conversations)", int((vc >= 2).sum())),
        (
            "Part d'utilisateurs récurrents (%)",
            round(100 * (vc >= 2).sum() / vc.shape[0], 1) if vc.shape[0] else 0.0,
        ),
        ("Durée médiane d'une conversation initiée (s)", int(dur.median()) if len(dur) else 0),
        ("Nb messages médian par conversation initiée", int(msg.median()) if len(msg) else 0),
        ("Tickets HubSpot créés depuis une conversation initiée", nb_tickets),
        (
            "Part de conversations débouchant sur un ticket (%)",
            round(100 * nb_tickets / len(df_init), 1) if len(df_init) else 0.0,
        ),
        ("Conversations initiées évaluées (LLM)", int(len(df_eval))),
        ("Satisfaction LLM globale (Sat / Sat+Insat) (%)", taux_sat),
    ]
    return pd.DataFrame(lignes, columns=["Indicateur", "Valeur"])


def build_methodologie(df_fse: pd.DataFrame, df_init: pd.DataFrame) -> pd.DataFrame:
    texte = [
        "Périmètre",
        "  - Source : export Yelda data/Affid/yelda/yelda.xlsx (une ligne = une conversation).",
        "  - Filtre : URL d'origine commençant par https://fse.stellair.fr (chatbot FSE Stellair).",
        "",
        "Définition d'une conversation « initiée »",
        "  - Critère : colonne 'Nombre de messages' >= 2.",
        "  - Justification : lorsque le widget est ouvert sans interaction, le bot envoie",
        "    uniquement son message de bienvenue (Parcours = 'welcome', nb_messages = 1,",
        "    durée quasi nulle). Dès qu'une question utilisateur est posée, le bot y répond,",
        "    donc le nombre de messages passe à 2 au minimum.",
        "  - Vérification sur le jeu de données : cohérence parfaite entre 'nb_messages >= 2' et",
        "    'Parcours != welcome'.",
        "",
        "Évaluation LLM des conversations (colonnes 'Évaluation LLM' et 'Score LLM')",
        "  - NE PROVIENT PAS de Yelda. Produite par un script Python interne qui lit le",
        "    champ 'Contenu de la conversation' (messages bot + utilisateur) et appelle Claude",
        "    (Anthropic) avec un prompt d'évaluation dédié, sans utiliser le 'Parcours' ni les",
        "    'Intentions' détectées.",
        "  - Le LLM renvoie un score entier de 0 à 5 + une raison courte :",
        "      5 / 4 → Satisfait                (problème résolu, réponse utile, ou clarification pertinente)",
        "      3    → À revoir                  (réponse partielle, ambiguë, sans proposition de suite)",
        "      2 / 1 → Insatisfait              (réponse inadéquate, frustration, échec)",
        "      0    → Non évaluable             (conversation vide ou sans question utilisateur)",
        "  - Les conversations où l'utilisateur n'a posé aucune question (pas de message 'user')",
        "    sont automatiquement marquées 'Non évaluable' (score 0), sans appel au LLM.",
        "  - Résultats stockés dans evaluation_yelda.json (cache incrémental) puis reportés",
        "    dans les colonnes 'Évaluation LLM' et 'Score LLM' de yelda.xlsx.",
        "",
        "Autres indicateurs",
        "  - Utilisateurs uniques : contacts distincts via le slot persistant hubspot_id_slot.",
        "  - Taux de satisfaction LLM = Satisfait / (Satisfait + Insatisfait) ; les 'À revoir'",
        "    et 'Non évaluable' sont exclus du dénominateur pour lisibilité (volumes faibles).",
        "  - Ticket créé : Parcours contient 'creation_ticket_hubspot'.",
        "  - Intentions : liste séparée par ';' dans la colonne 'Intentions'. Le tri compte",
        "    chaque occurrence unique par conversation.",
        "",
        "Chiffres clés (mise à jour à chaque export)",
        f"  - Ouvertures du widget : {len(df_fse)}",
        f"  - Conversations initiées : {len(df_init)} "
        f"({round(100*len(df_init)/len(df_fse),1) if len(df_fse) else 0} %)",
    ]
    return pd.DataFrame({"Méthodologie": texte})


def export_yelda_comex_workbook(out_xlsx: Path) -> Dict[str, pd.DataFrame]:
    df = load_yelda_data()
    if df is None or df.empty:
        raise SystemExit("Fichier Yelda introuvable ou vide.")
    df_fse = filter_yelda_stellair(df)
    df_init = filter_conversations_initiees(df_fse)

    sheets: Dict[str, pd.DataFrame] = {
        "Synthese": build_synthese(df_fse, df_init),
        "Volumetrie_hebdo": build_volumetrie_hebdo(df_init),
        "Volumetrie_mensuelle": build_volumetrie_mensuelle(df_init),
        "Utilisateurs": build_utilisateurs(df_init),
        "Satisfaction_LLM_mensuel": build_satisfaction_llm_mensuel(df_init),
        "Deflexion_ticket": build_deflexion_ticket(df_init),
        "Top_intentions": build_top_intentions(df_init),
        "Methodologie": build_methodologie(df_fse, df_init),
    }

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        for name, dfi in sheets.items():
            dfi.to_excel(xw, sheet_name=name, index=False)

    _format_workbook(out_xlsx, sheets)
    return sheets


def _format_workbook(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    """Mise en forme légère + 2 graphiques (volumétrie, satisfaction LLM)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for name, df in sheets.items():
        ws = wb[name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for c in col:
                v = "" if c.value is None else str(c.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Graphique volumétrie hebdo
    vol_h = sheets.get("Volumetrie_hebdo", pd.DataFrame())
    if not vol_h.empty:
        ws = wb["Volumetrie_hebdo"]
        ch = BarChart()
        ch.title = "Conversations initiées par semaine ISO"
        ch.y_axis.title = "Nb conversations"
        ch.x_axis.title = "Semaine"
        col_conv = list(vol_h.columns).index("Conversations initiées") + 1
        data = Reference(ws, min_col=col_conv, min_row=1, max_row=1 + len(vol_h), max_col=col_conv)
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(vol_h))
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 9
        ch.width = 22
        ws.add_chart(ch, "H2")

    # Graphique volumétrie mensuelle
    vol = sheets["Volumetrie_mensuelle"]
    if not vol.empty:
        ws = wb["Volumetrie_mensuelle"]
        ch = BarChart()
        ch.title = "Conversations initiées par mois"
        ch.y_axis.title = "Nb conversations"
        ch.x_axis.title = "Mois"
        data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(vol), max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(vol))
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 9
        ch.width = 18
        ws.add_chart(ch, "G2")

    # Graphique satisfaction LLM
    sat = sheets["Satisfaction_LLM_mensuel"]
    if not sat.empty:
        ws = wb["Satisfaction_LLM_mensuel"]
        ch = LineChart()
        ch.title = "Taux de satisfaction LLM (%) par mois"
        ch.y_axis.title = "% Satisfait / (Sat+Insat)"
        ch.x_axis.title = "Mois"
        col_taux = sat.columns.get_loc("Taux satisfaction LLM (%)") + 1
        data = Reference(ws, min_col=col_taux, min_row=1, max_row=1 + len(sat), max_col=col_taux)
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(sat))
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 9
        ch.width = 18
        ws.add_chart(ch, f"H2")

    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(description="Classeur Excel COMEX — utilisation du chatbot Yelda.")
    default_out = _ROOT / "data/Affid/analyse_appels_tickets/yelda_comex_analyse_utilisation.xlsx"
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=default_out,
        help=f"Fichier .xlsx de sortie (défaut : {default_out.relative_to(_ROOT)})",
    )
    args = p.parse_args()

    sheets = export_yelda_comex_workbook(args.output)
    print(f"Écrit : {args.output.resolve()}")
    print(f"Feuilles : {', '.join(sheets.keys())}")


if __name__ == "__main__":
    main()
