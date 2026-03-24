"""
Analyse et catégorisation des transcriptions d'appels - Support Hardware (01 89 74 42 59)

Service rattaché récemment : pharmaciens et laboratoires avec borne de mise à jour carte vitale.
Extrait : sujets, demandes, secteur (support, administratif, logistique)
"""

import os
import re
import json
from collections import defaultdict
from datetime import datetime
from typing import Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv

# Charger les variables d'environnement
if os.path.exists('.env'):
    load_dotenv('.env')
elif os.path.exists('config.env'):
    load_dotenv('config.env')

API_ID = os.getenv('AIRCALL_API_ID')
API_TOKEN = os.getenv('AIRCALL_API_TOKEN')
BASE_URL = 'https://api.aircall.io/v1'

# Numéro Support Hardware - borne carte vitale pharmaciens/laboratoires
SUPPORT_HARDWARE_NUMBER_ID = 1172775
SUPPORT_HARDWARE_DIGITS = "1 89 74 42 59"

# ============ RÈGLES DE CATÉGORISATION ============
# Adaptées au contexte : borne, lecteur, carte vitale, pharmaciens, laboratoires

# Secteurs avec mots-clés et poids
SECTEUR_KEYWORDS = {
    "support": [
        "problème", "ne marche pas", "ne fonctionne pas", "erreur", "bug", "dysfonctionnement",
        "aide", "assistance", "configuration", "paramétrage", "dépannage", "résolution",
        "carte vitale", "lecteur", "borne", "passer la carte", "mise à jour", "tmaj",
        "droits amo", "droits fermés", "tiers payant", "complémentaire santé", "css", "msa",
        "sans contact", "e-carte vital", "écarte vital", "lecture", "ne lit pas", "ne passe pas",
        "bloqué", "planté", "redémarrer", "redémarrage", "écran", "affichage",
    ],
    "administratif": [
        "facture", "facturation", "contrat", "abonnement", "paiement", "commande",
        "devis", "tarif", "prix", "récapitulatif", "document", "attestation",
        "compte", "identifiants", "login", "mot de passe", "création de compte",
        "résiliation", "renouvellement", "engagement",
    ],
    "logistique": [
        "livraison", "envoi", "réception", "colis", "expédition",
        "sav", "retour", "réparation", "remplacement", "échange",
        "installation", "mise en place", "déploiement", "intervention",
        "enlever", "remettre", "retirer", "déposer", "récupérer",
        "lecteur", "borne", "équipement", "matériel", "appareil",
        "numéro de série", "référence", "garantie",
    ],
}

# Sujets thématiques (pour affiner l'analyse)
SUJET_KEYWORDS = {
    "borne_carte_vitale": ["borne", "carte vitale", "lecteur", "mise à jour", "tmaj", "passer la carte"],
    "droits_patients": ["droits amo", "droits fermés", "tiers payant", "complémentaire", "css", "msa", "am0"],
    "technique_lecteur": ["sans contact", "e-carte vital", "lecture", "ne lit pas", "configuration"],
    "facturation_administratif": ["facture", "facturation", "contrat", "abonnement", "devis"],
    "logistique_equipement": ["livraison", "installation", "remplacement", "enlever", "remettre", "sav"],
    "compte_identifiants": ["compte", "login", "identifiants", "mot de passe", "création"],
}

# Types de demandes
DEMANDE_KEYWORDS = {
    "depannage": ["problème", "ne marche", "erreur", "aide", "assistance", "dépannage"],
    "information": ["question", "savoir", "comment", "information", "renseignement"],
    "action": ["enlever", "remettre", "installer", "remplacer", "envoyer", "réparer"],
    "document": ["facture", "document", "devis", "attestation", "récapitulatif"],
    "accompagnement": ["accompagnement", "formation", "prise en main", "mettre en place"],
}

# Sujets des questions support (granularité pour les demandes support)
SUJET_QUESTION_SUPPORT = {
    "configuration_parametrage": [
        "configurer", "paramétrer", "paramétrage", "configuration", "réglage",
        "mise en route", "mise en service", "activer", "activation",
    ],
    "panne_dysfonctionnement": [
        "ne marche pas", "ne fonctionne pas", "bloqué", "planté", "plus rien",
        "dysfonctionnement", "panne", "cassé", "ne s'allume pas", "ne démarre pas",
    ],
    "erreur_code": [
        "erreur", "code erreur", "message d'erreur", "erreur 11", "erreur 1121",
        "ne reconnaît pas", "ne détecte pas", "ne lit pas", "ne passe pas",
    ],
    "installation_reinstallation": [
        "installer", "réinstaller", "installation", "changement d'ordinateur",
        "nouveau poste", "nouveau pc", "brancher", "connecter",
    ],
    "mise_a_jour_tmaj": [
        "mise à jour", "tmaj", "mettre à jour", "cartes vitales", "passer la carte",
        "ne fait pas de mise à jour", "mise à jour ne fonctionne",
    ],
    "lecture_carte": [
        "lecture", "lire", "carte vitale", "sans contact", "e-carte vital",
        "ne lit plus", "ne reconnaît plus les cartes",
    ],
    "droits_patients": [
        "droits amo", "droits fermés", "tiers payant", "am0", "complémentaire",
        "css", "msa", "support de droit",
    ],
    "remplacement_sav": [
        "remplacer", "remplacement", "sav", "réparation", "bouton cassé",
        "en panne", "défaillant",
    ],
    "accompagnement_formation": [
        "accompagnement", "formation", "prise en main", "comment faire",
        "savoir utiliser", "paramétrer les lecteurs",
    ],
}


def normalize_text(text: str) -> str:
    """Normalise le texte pour la recherche (minuscules, accents approximatifs)"""
    if not text:
        return ""
    text = text.lower().strip()
    # Remplacer quelques variations courantes
    text = re.sub(r'\s+', ' ', text)
    return text


def extract_full_transcript(transcription: dict) -> str:
    """Extrait le texte complet de la transcription (priorité aux interventions externes)"""
    content = transcription.get("content", {}) or {}
    utterances = content.get("utterances", [])
    parts = []
    for u in utterances:
        text = u.get("text", "").strip()
        if text:
            parts.append(text)
    return " ".join(parts)


def categorize_by_keywords(text: str, keyword_sets: dict) -> dict:
    """
    Catégorise le texte selon des ensembles de mots-clés.
    Retourne un dict {catégorie: score} où score = nombre de mots trouvés
    """
    normalized = normalize_text(text)
    results = {}
    for category, keywords in keyword_sets.items():
        score = 0
        matched = []
        for kw in keywords:
            if kw.lower() in normalized:
                score += 1
                matched.append(kw)
        if score > 0:
            results[category] = {"score": score, "matched": matched}
    return results


def export_to_excel(synthese: dict, filepath: str) -> None:
    """Exporte les résultats vers un fichier Excel."""
    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # --- Feuille 1 : Détail des appels ---
    ws_detail = wb.active
    ws_detail.title = "Détail des appels"
    cols_detail = [
        "ID Appel", "Date", "Durée (s)", "Secteur", "Confiance",
        "Sujet", "Mots clés", "Sujets", "Demandes", "Question client"
    ]
    for col, header in enumerate(cols_detail, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws_detail.row_dimensions[1].height = 25

    details = synthese.get("details", [])
    for row_idx, r in enumerate(details, 2):
        a = r.get("analysis", {})
        sujet_appel = a.get("sujet_appel", a.get("sujet_question_client", {}))
        sujet = ", ".join(s.replace("_", " ").title() for s in sujet_appel.keys())
        mots_cles = ", ".join(a.get("mots_cles", []))
        sujets = ", ".join(a.get("sujets", {}).keys())
        demandes = ", ".join(a.get("demandes", {}).keys())
        question = (a.get("resume_demande") or "")[:2000]

        ws_detail.cell(row=row_idx, column=1, value=r.get("call_id"))
        ws_detail.cell(row=row_idx, column=2, value=r.get("date") or "N/A")
        ws_detail.cell(row=row_idx, column=3, value=r.get("duration", 0))
        ws_detail.cell(row=row_idx, column=4, value=a.get("secteur_principal", ""))
        ws_detail.cell(row=row_idx, column=5, value=a.get("confiance", 0))
        ws_detail.cell(row=row_idx, column=6, value=sujet or "-")
        ws_detail.cell(row=row_idx, column=7, value=mots_cles or "-")
        ws_detail.cell(row=row_idx, column=8, value=sujets or "-")
        ws_detail.cell(row=row_idx, column=9, value=demandes or "-")
        ws_detail.cell(row=row_idx, column=10, value=question)
        ws_detail.cell(row=row_idx, column=10).alignment = Alignment(wrap_text=True)

    for col in range(1, 11):
        ws_detail.column_dimensions[get_column_letter(col)].width = max(12, min(50, 18 if col < 10 else 60))

    # --- Feuille 2 : Synthèse ---
    ws_synthese = wb.create_sheet("Synthèse", 1)
    ws_synthese.cell(row=1, column=1, value="Synthèse - Support Hardware (01 89 74 42 59)")
    ws_synthese.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_synthese.merge_cells("A1:B1")

    row = 3
    periode = synthese.get("periode", {})
    ws_synthese.cell(row=row, column=1, value="Période analysée")
    ws_synthese.cell(row=row, column=1).font = header_font
    row += 1
    ws_synthese.cell(row=row, column=1, value="Nombre d'appels")
    ws_synthese.cell(row=row, column=2, value=periode.get("nb_appels", 0))
    row += 1
    ws_synthese.cell(row=row, column=1, value="Transcriptions analysées")
    ws_synthese.cell(row=row, column=2, value=periode.get("nb_transcriptions", 0))
    row += 2

    ws_synthese.cell(row=row, column=1, value="Répartition par secteur")
    ws_synthese.cell(row=row, column=1).font = header_font
    row += 1
    for secteur, count in synthese.get("repartition_secteurs", {}).items():
        ws_synthese.cell(row=row, column=1, value=secteur)
        ws_synthese.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    ws_synthese.cell(row=row, column=1, value="Sujets fréquents")
    ws_synthese.cell(row=row, column=1).font = header_font
    row += 1
    for sujet, count in list(synthese.get("sujets_frequents", {}).items())[:15]:
        ws_synthese.cell(row=row, column=1, value=sujet.replace("_", " ").title())
        ws_synthese.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    ws_synthese.cell(row=row, column=1, value="Sujets des appels")
    ws_synthese.cell(row=row, column=1).font = header_font
    row += 1
    for sq, count in synthese.get("sujets_questions_support", {}).items():
        ws_synthese.cell(row=row, column=1, value=sq.replace("_", " ").title())
        ws_synthese.cell(row=row, column=2, value=count)
        row += 1

    ws_synthese.column_dimensions["A"].width = 35
    ws_synthese.column_dimensions["B"].width = 15

    wb.save(filepath)


def get_primary_secteur(secteur_scores: dict) -> Tuple[str, float]:
    """Retourne le secteur principal et son score relatif"""
    if not secteur_scores:
        return ("non_classé", 0.0)
    best = max(secteur_scores.items(), key=lambda x: x[1]["score"])
    total = sum(s["score"] for s in secteur_scores.values())
    return (best[0], best[1]["score"] / max(total, 1))


def analyze_transcription(transcription: dict) -> dict:
    """
    Analyse une transcription et retourne les catégorisations.
    """
    full_text = extract_full_transcript(transcription)
    if not full_text or len(full_text) < 20:
        return {
            "secteur_principal": "non_classé",
            "secteurs": {},
            "sujets": {},
            "demandes": {},
            "resume_demande": "",
            "sujet_appel": {},
            "mots_cles": [],
            "confiance": 0.0,
        }

    # Catégorisation
    secteurs = categorize_by_keywords(full_text, SECTEUR_KEYWORDS)
    sujets = categorize_by_keywords(full_text, SUJET_KEYWORDS)
    demandes = categorize_by_keywords(full_text, DEMANDE_KEYWORDS)

    secteur_principal, confiance = get_primary_secteur(secteurs)

    # Résumé de la demande (extrait des premières interventions externes = client)
    content = transcription.get("content", {}) or {}
    utterances = content.get("utterances", [])
    external_texts = [
        u.get("text", "") for u in utterances
        if u.get("participant_type") == "external" and u.get("text")
    ]
    resume_demande = " ".join(external_texts[:3])[:500] if external_texts else full_text[:500]

    # Sujet de l'appel (pour TOUS les appels) - catégorisation fine
    sujet_appel_raw = categorize_by_keywords(
        resume_demande or full_text, SUJET_QUESTION_SUPPORT
    )
    sujet_appel = {k: v["matched"] for k, v in sujet_appel_raw.items() if v["matched"]}

    # Mots clés : tous les termes détectés (sujets, demandes, secteurs, sujet_appel)
    mots_cles = set()
    for d in [sujets, demandes, secteurs, sujet_appel_raw]:
        for v in d.values():
            if isinstance(v, dict) and "matched" in v:
                mots_cles.update(v["matched"])
    mots_cles = sorted(mots_cles)

    return {
        "secteur_principal": secteur_principal,
        "secteurs": {k: v["score"] for k, v in secteurs.items()},
        "sujets": {k: v["matched"] for k, v in sujets.items() if v["matched"]},
        "demandes": {k: v["matched"] for k, v in demandes.items() if v["matched"]},
        "resume_demande": resume_demande.strip(),
        "sujet_appel": sujet_appel,  # Sujet pour TOUS les appels
        "mots_cles": mots_cles,  # Mots clés détectés
        "confiance": round(confiance, 2),
    }


class AircallAnalyzer:
    """Analyseur des appels Support Hardware"""

    def __init__(self, api_id: str, api_token: str):
        self.auth = HTTPBasicAuth(api_id, api_token)
        self.base_url = BASE_URL

    def _request(self, endpoint: str, params: Optional[dict] = None) -> Optional[dict]:
        url = f"{self.base_url}/{endpoint}"
        try:
            r = requests.get(url, auth=self.auth, params=params or {}, timeout=30)
            r.raise_for_status()
            return r.json()
        except requests.RequestException as e:
            print(f"   Erreur API: {e}")
            return None

    def get_calls_support_hardware(
        self, limit: int = 100, direction: str = "inbound"
    ) -> list:
        """Récupère les appels inbound du numéro Support Hardware"""
        all_calls = []
        page = 1
        per_page = 50

        while len(all_calls) < limit:
            data = self._request(
                "calls",
                params={
                    "per_page": per_page,
                    "page": page,
                    "order": "desc",
                },
            )
            if not data or "calls" not in data:
                break

            calls = data["calls"]
            for call in calls:
                num = call.get("number") or {}
                if isinstance(num, dict) and num.get("id") == SUPPORT_HARDWARE_NUMBER_ID:
                    if direction is None or call.get("direction") == direction:
                        all_calls.append(call)
                        if len(all_calls) >= limit:
                            break

            if len(calls) < per_page:
                break
            page += 1
            if len(all_calls) >= limit:
                break

        return all_calls[:limit]

    def get_transcription(self, call_id: int) -> Optional[dict]:
        """Récupère la transcription d'un appel"""
        data = self._request(f"calls/{call_id}/transcription")
        return data.get("transcription") if data else None

    def analyze_support_hardware_calls(
        self, limit: int = 50, save_json: bool = True, save_excel: bool = True
    ) -> dict:
        """
        Pipeline complet : récupère les appels, transcriptions, et analyse.
        """
        print("=" * 70)
        print("ANALYSE DES APPELS - Support Hardware (01 89 74 42 59)")
        print("=" * 70)
        print()

        # 1. Récupérer les appels inbound
        print("1️⃣  Récupération des appels inbound...")
        calls = self.get_calls_support_hardware(limit=limit, direction="inbound")
        print(f"   → {len(calls)} appel(s) inbound trouvé(s)")
        print()

        if not calls:
            return {"calls": [], "synthese": {}}

        # 2. Récupérer les transcriptions et analyser
        print("2️⃣  Récupération des transcriptions et analyse...")
        results = []
        stats_secteurs = defaultdict(int)
        stats_sujets = defaultdict(int)
        stats_sujet_question_support = defaultdict(int)

        for i, call in enumerate(calls, 1):
            call_id = call.get("id")
            started_at = call.get("started_at")
            duration = call.get("duration", 0)

            transcription = self.get_transcription(call_id)
            if not transcription:
                print(f"   ⚠ Appel {call_id}: pas de transcription")
                continue

            analysis = analyze_transcription(transcription)
            stats_secteurs[analysis["secteur_principal"]] += 1
            for sujet in analysis.get("sujets", {}).keys():
                stats_sujets[sujet] += 1
            for sq in analysis.get("sujet_appel", {}).keys():
                stats_sujet_question_support[sq] += 1

            results.append({
                "call_id": call_id,
                "date": datetime.fromtimestamp(started_at).strftime("%Y-%m-%d %H:%M") if started_at else "N/A",
                "duration": duration,
                "analysis": analysis,
            })
            print(f"   ✓ Appel {call_id}: {analysis['secteur_principal']}")

        print()
        print(f"   → {len(results)} transcription(s) analysée(s)")
        print()

        # 3. Synthèse
        synthese = {
            "periode": {
                "nb_appels": len(calls),
                "nb_transcriptions": len(results),
            },
            "repartition_secteurs": dict(stats_secteurs),
            "sujets_frequents": dict(
                sorted(stats_sujets.items(), key=lambda x: -x[1])[:10]
            ),
            "sujets_questions_support": dict(
                sorted(stats_sujet_question_support.items(), key=lambda x: -x[1])
            ),
            "details": results,
        }

        # 4. Affichage
        print("3️⃣  SYNTHÈSE")
        print("-" * 50)
        print("\n📊 Répartition par secteur:")
        for secteur, count in sorted(
            stats_secteurs.items(), key=lambda x: -x[1]
        ):
            pct = 100 * count / len(results) if results else 0
            print(f"   • {secteur}: {count} ({pct:.0f}%)")

        print("\n📌 Sujets les plus fréquents:")
        for sujet, count in list(
            sorted(stats_sujets.items(), key=lambda x: -x[1])
        )[:8]:
            print(f"   • {sujet}: {count}")

        if stats_sujet_question_support:
            print("\n❓ Sujets des appels (tous secteurs):")
            for sq, count in sorted(
                stats_sujet_question_support.items(), key=lambda x: -x[1]
            ):
                label = sq.replace("_", " ").title()
                print(f"   • {label}: {count}")

        print("\n📝 Détail des appels analysés:")
        print("-" * 50)
        for r in results:
            a = r["analysis"]
            print(f"\n   Appel #{r['call_id']} | {r['date']} | {r['duration']}s")
            print(f"   Secteur: {a['secteur_principal']} (confiance: {a['confiance']})")
            if a.get("sujet_appel"):
                sq_labels = [s.replace("_", " ").title() for s in a["sujet_appel"].keys()]
                print(f"   Sujet: {', '.join(sq_labels)}")
            if a.get("mots_cles"):
                print(f"   Mots clés: {', '.join(a['mots_cles'][:15])}{'...' if len(a['mots_cles']) > 15 else ''}")
            if a.get("sujets"):
                print(f"   Sujets: {', '.join(a['sujets'].keys())}")
            if a.get("demandes"):
                print(f"   Demandes: {', '.join(a['demandes'].keys())}")
            if a.get("resume_demande"):
                resume = a["resume_demande"][:200] + "..." if len(a["resume_demande"]) > 200 else a["resume_demande"]
                print(f"   Question client: {resume}")

        # 5. Sauvegarde JSON
        if save_json:
            filename = f"analyse_support_hardware_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(synthese, f, indent=2, ensure_ascii=False)
            print()
            print(f"💾 Résultats sauvegardés: {filename}")

        # 6. Sauvegarde Excel
        if save_excel and results:
            excel_filename = f"analyse_support_hardware_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            try:
                export_to_excel(synthese, excel_filename)
                print(f"📊 Export Excel: {excel_filename}")
            except Exception as e:
                print(f"   ⚠ Erreur export Excel: {e}")

        print()
        print("=" * 70)
        return synthese


def main():
    if not API_ID or not API_TOKEN:
        print("❌ Configurez AIRCALL_API_ID et AIRCALL_API_TOKEN dans config.env")
        return

    analyzer = AircallAnalyzer(API_ID, API_TOKEN)
    analyzer.analyze_support_hardware_calls(limit=50)


if __name__ == "__main__":
    main()
